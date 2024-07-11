import tkinter as tk
from tkinter import messagebox, simpledialog
from openpyxl import Workbook, load_workbook
from datetime import datetime

class Produto:
    def __init__(self, data, cliente, codigo, loja, quantidade, nome, valor_dolar_unidade, porcentagem_lucro):
        self.data = data
        self.cliente = cliente
        self.codigo = codigo
        self.loja = loja
        self.quantidade = quantidade
        self.nome = nome
        self.valor_dolar_unidade = valor_dolar_unidade
        self.porcentagem_lucro = porcentagem_lucro

    def calcular_valor_total_dolar(self):
        return self.quantidade * self.valor_dolar_unidade

    def calcular_valor_total_reais(self, taxa_conversao):
        valor_total_dolar = self.calcular_valor_total_dolar()
        valor_total_reais = valor_total_dolar * taxa_conversao
        return valor_total_reais

    def calcular_lucro_dolar(self):
        valor_total_dolar = self.calcular_valor_total_dolar()
        lucro_dolar = valor_total_dolar * (self.porcentagem_lucro / 100)
        return lucro_dolar

    def calcular_lucro_reais(self, taxa_conversao):
        lucro_dolar = self.calcular_lucro_dolar()
        lucro_reais = lucro_dolar * taxa_conversao
        return lucro_reais

    def calcular_lucro_por_unidade_reais(self, taxa_conversao):
        lucro_por_unidade_reais = self.calcular_lucro_reais(taxa_conversao) / self.quantidade
        return lucro_por_unidade_reais


class BancoDeDadosProdutos:
    def __init__(self):
        self.produtos = {}
        self.interface = None  # Referência à interface
        self.taxa_conversao = 5.0  # padrão da taxa de conversão (1 dólar = 5 reais)
        self.filename = self.get_nome_planilha()
        self.carregar_produtos()
        

    def set_interface(self, interface):
        self.interface = interface

    def get_nome_planilha(self):
        now = datetime.now()
        nome_mes_ano = now.strftime("planilha_%m_%Y.xlsx")
        return nome_mes_ano

    def carregar_produtos(self):
        try:
            wb = load_workbook(self.filename)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = row[0]
                codigo = row[2]
                cliente = row[1]
                loja = row[3]
                quantidade = row[4]
                nome = row[5]
                valor_dolar_unidade = row[6]
                porcentagem_lucro = row[11]
                produto = Produto(data, cliente, codigo, loja, quantidade, nome, valor_dolar_unidade, porcentagem_lucro)
                self.produtos[codigo] = produto
        except FileNotFoundError:
            # Se o arquivo não existir, cria um novo durante a primeira execução
            pass

    def adicionar_produto(self, data, cliente, codigo, loja, quantidade, nome, valor_dolar_unidade, porcentagem_lucro):
        produto = Produto(data, cliente, codigo, loja, quantidade, nome, valor_dolar_unidade, porcentagem_lucro)
        self.produtos[codigo] = produto
        self.salvar_em_excel(produto)
        return True

    def salvar_em_excel(self, produto):
        filename = self.get_nome_planilha()
        try:
            wb = load_workbook(filename)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["data", "Cliente", "Código", "Loja", "Quantidade", "Nome do Produto", "Valor Unitário (Dólar)",
                    "Valor Total (Dólar)", "Taxa de Conversão", "Valor Unitário (Real)", "Valor Total (Real)",
                    "Porcentagem de Lucro", "Lucro por Unidade (Real)", "Lucro Total (Real)"])

        valor_total_dolar = produto.calcular_valor_total_dolar()
        valor_total_reais = valor_total_dolar * self.taxa_conversao
        valor_unitario_reais = produto.valor_dolar_unidade * self.taxa_conversao
        lucro_total_reais = produto.calcular_lucro_reais(self.taxa_conversao)
        lucro_por_unidade_reais = produto.calcular_lucro_por_unidade_reais(self.taxa_conversao)

        ws.append([produto.data.strftime("%Y-%m-%d"), produto.cliente, produto.codigo, produto.loja, produto.quantidade, produto.nome,
                f'{produto.valor_dolar_unidade:.2f}', f'{valor_total_dolar:.2f}',
                f'{self.taxa_conversao:.2f}', f'{valor_unitario_reais:.2f}',
                f'{valor_total_reais:.2f}', produto.porcentagem_lucro, lucro_por_unidade_reais, f'{lucro_total_reais:.2f}'])

        try:
            wb.save(filename)
            messagebox.showinfo("Sucesso", "Produtos salvos em arquivo Excel.")
        except PermissionError:
            messagebox.showerror("Erro", "Não foi possível salvar a planilha. Verifique as permissões de escrita.")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao salvar a planilha: {e}")


    def atualizar_cotacao(self):
        try:
            nova_cotacao = float(simpledialog.askstring("Atualizar Cotação", "Digite a nova cotação do Dólar (R$):"))
            self.taxa_conversao = nova_cotacao
            # Atualiza o texto do rótulo na interface
            self.interface.label_cotacao.config(text=f"Cotação do Dólar (R$): {self.taxa_conversao:.2f}")
            messagebox.showinfo("Sucesso", f"Cotação do Dólar atualizada para R$ {nova_cotacao:.2f}.")
        except ValueError:
            messagebox.showerror("Erro", "Digite um valor numérico válido para a cotação do Dólar.")

    def calcular_total_mes(self):
        filename = self.get_nome_planilha()
        try:
            wb = load_workbook(filename)
            ws = wb.active
            total_dolares = 0.0
            total_reais = 0.0
            total_lucro_mes = 0.0
            for row in ws.iter_rows(min_row=2, values_only=True):
                valor_total_dolar = row[7]
                valor_total_reais = row[10]
                lucro_reais = row[13]

                # Verificação para garantir que os valores sejam numéricos
                if valor_total_dolar is not None:
                    try:
                        valor_total_dolar = float(valor_total_dolar)
                        total_dolares += valor_total_dolar
                    except ValueError:
                        pass
                if valor_total_reais is not None:
                    try:
                        valor_total_reais = float(valor_total_reais)
                        total_reais += valor_total_reais
                    except ValueError:
                        pass
                if lucro_reais is not None:
                    try:
                        lucro_reais = float(lucro_reais)
                        total_lucro_mes += lucro_reais
                    except ValueError:
                        pass
                
                print(f"Row: {row}")
                print(f"Valor Total Dólar: {valor_total_dolar}, Valor Total Reais: {valor_total_reais}, Lucro Reais: {lucro_reais}")

            messagebox.showinfo("Total do Mês", f"Total gasto no mês:\n\nDólares: ${total_dolares:.2f}\nReais: R${total_reais:.2f}\nTotal de lucro no mês: R${total_lucro_mes:.2f}")
        except FileNotFoundError:
            messagebox.showerror("Erro", "Nenhuma planilha encontrada para calcular o total do mês.")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao calcular o total do mês: {e}")


class InterfaceProdutos:
    def __init__(self, root):
        self.root = root
        self.root.title("Gerenciador de Produtos")
        self.banco_produtos = BancoDeDadosProdutos()
        self.banco_produtos.set_interface(self)

        # Componentes da interface
        self.label_cliente = tk.Label(root, text="Cliente:")
        self.entry_cliente = tk.Entry(root)

        self.label_codigo = tk.Label(root, text="Código do Produto:")
        self.entry_codigo = tk.Entry(root)

        self.radio_var = tk.StringVar()  # Variável para armazenar a seleção do usuário
        self.radio_var.set("unidade")  # Configuração padrão

        self.radio_por_unidade = tk.Radiobutton(root, text="Por Unidade", variable=self.radio_var, value="unidade", command=self.atualizar_interface)
        self.radio_por_porcentagem = tk.Radiobutton(root, text="Por Porcentagem", variable=self.radio_var, value="porcentagem", command=self.atualizar_interface)

        self.label_loja = tk.Label(root, text="Loja:")
        self.entry_loja = tk.Entry(root)

        self.label_quantidade = tk.Label(root, text="Quantidade:")
        self.entry_quantidade = tk.Entry(root)

        self.label_nome = tk.Label(root, text="Nome do Produto:")
        self.entry_nome = tk.Entry(root)

        self.label_valor_dolar_unidade = tk.Label(root, text="Valor em Dólar por Unidade:")
        self.entry_valor_dolar_unidade = tk.Entry(root)

        self.label_porcentagem_lucro = tk.Label(root, text="Porcentagem de Lucro (%):")
        self.entry_porcentagem_lucro = tk.Entry(root)

        self.label_lucro_por_unidade = tk.Label(root, text="Lucro por Unidade (R$):")
        self.entry_lucro_por_unidade = tk.Entry(root)

        self.botao_adicionar = tk.Button(root, text="Adicionar Produto", command=self.adicionar_produto)
        self.botao_atualizar_cotacao = tk.Button(root, text="Atualizar Cotação", command=self.banco_produtos.atualizar_cotacao)
        self.botao_calcular_total_mes = tk.Button(root, text="Calcular Total do Mês", command=self.banco_produtos.calcular_total_mes)

        self.label_cotacao = tk.Label(root, text=f"Cotação do Dólar (R$): {self.banco_produtos.taxa_conversao:.2f}")

        # Layout
        self.label_cliente.grid(row=0, column=0, padx=10, pady=5, sticky="w")
        self.entry_cliente.grid(row=0, column=1, padx=10, pady=5)

        self.label_codigo.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        self.entry_codigo.grid(row=1, column=1, padx=10, pady=5)

        self.radio_por_unidade.grid(row=7, column=0, padx=10, pady=5)
        self.radio_por_porcentagem.grid(row=7, column=1, padx=10, pady=5)

        self.label_loja.grid(row=2, column=0, padx=10, pady=5, sticky="w")
        self.entry_loja.grid(row=2, column=1, padx=10, pady=5)

        self.label_quantidade.grid(row=3, column=0, padx=10, pady=5, sticky="w")
        self.entry_quantidade.grid(row=3, column=1, padx=10, pady=5)

        self.label_nome.grid(row=4, column=0, padx=10, pady=5, sticky="w")
        self.entry_nome.grid(row=4, column=1, padx=10, pady=5)

        self.label_valor_dolar_unidade.grid(row=5, column=0, padx=10, pady=5, sticky="w")
        self.entry_valor_dolar_unidade.grid(row=5, column=1, padx=10, pady=5)

        self.label_porcentagem_lucro.grid(row=6, column=0, padx=10, pady=5, sticky="w")
        self.entry_porcentagem_lucro.grid(row=6, column=1, padx=10, pady=5)

        self.label_lucro_por_unidade.grid(row=8, column=0, padx=10, pady=5, sticky="w")
        self.entry_lucro_por_unidade.grid(row=8, column=1, padx=10, pady=5)

        self.botao_adicionar.grid(row=9, column=0, columnspan=2, padx=10, pady=10, sticky="we")
        self.botao_atualizar_cotacao.grid(row=10, column=0, columnspan=2, padx=10, pady=5, sticky="we")
        self.botao_calcular_total_mes.grid(row=11, column=0, columnspan=2, padx=10, pady=5, sticky="we")

        self.label_cotacao.grid(row=12, column=0, columnspan=2, padx=10, pady=5, sticky="w")
        

        # Atualizar a interface de acordo com a seleção do usuário
        self.atualizar_interface()
    

    def adicionar_produto(self):
        data = datetime.now()
        cliente = self.entry_cliente.get()
        codigo = self.entry_codigo.get()
        loja = self.entry_loja.get()
        quantidade = self.entry_quantidade.get()
        nome = self.entry_nome.get()
        valor_dolar_unidade = self.entry_valor_dolar_unidade.get()

        # Verifica qual modo está selecionado e obtém os valores correspondentes
        if self.radio_var.get() == "unidade":
            lucro_por_unidade = self.entry_lucro_por_unidade.get()
            porcentagem_lucro = 0  # Será calculado mais tarde
        else:
            porcentagem_lucro = self.entry_porcentagem_lucro.get()
            lucro_por_unidade = 0  # Não usado nesse caso

        # Verifica se algum campo obrigatório está vazio
        if '' in (cliente, codigo, loja, quantidade, valor_dolar_unidade):
            messagebox.showerror("Erro", "Por favor, preencha todos os campos obrigatórios.")
            return

        try:
            # Converte os campos para os tipos adequados
            quantidade = int(quantidade)
            valor_dolar_unidade = float(valor_dolar_unidade)
            if lucro_por_unidade:
                lucro_por_unidade = float(lucro_por_unidade)
            if porcentagem_lucro:
                porcentagem_lucro = float(porcentagem_lucro)
        except ValueError:
            messagebox.showerror("Erro", "Por favor, insira valores numéricos válidos.")
            return

        # Calcula a porcentagem de lucro com base no lucro por unidade, se necessário
        if self.radio_var.get() == "unidade" and lucro_por_unidade:
            taxa_conversao = self.banco_produtos.taxa_conversao
            lucro_total_reais = lucro_por_unidade * quantidade
            lucro_total_dolar = lucro_total_reais / taxa_conversao
            valor_total_dolar = valor_dolar_unidade * quantidade
            porcentagem_lucro = (lucro_total_dolar / valor_total_dolar) * 100

        # Verifica se o código do produto já existe e, se sim, usa o nome existente
        if codigo in self.banco_produtos.produtos:
            nome = self.banco_produtos.produtos[codigo].nome
        else:
            nome = self.entry_nome.get()

        # Adiciona o produto ao banco de dados
        if self.banco_produtos.adicionar_produto(data, cliente, codigo, loja, quantidade, nome, valor_dolar_unidade, porcentagem_lucro):
            # Limpa os campos de entrada
            self.entry_cliente.delete(0, tk.END)
            self.entry_codigo.delete(0, tk.END)
            self.entry_loja.delete(0, tk.END)
            self.entry_quantidade.delete(0, tk.END)
            self.entry_nome.delete(0, tk.END)
            self.entry_valor_dolar_unidade.delete(0, tk.END)
            self.entry_porcentagem_lucro.delete(0, tk.END)
            self.entry_lucro_por_unidade.delete(0, tk.END)

    # Atualiza a interface para habilitar/desabilitar campos conforme a seleção do usuário
    def atualizar_interface(self):
        if self.radio_var.get() == "unidade":
            self.label_porcentagem_lucro.config(state="disabled")
            self.entry_porcentagem_lucro.config(state="disabled")
            self.label_lucro_por_unidade.config(state="normal")
            self.entry_lucro_por_unidade.config(state="normal")
        else:
            self.label_porcentagem_lucro.config(state="normal")
            self.entry_porcentagem_lucro.config(state="normal")
            self.label_lucro_por_unidade.config(state="disabled")
            self.entry_lucro_por_unidade.config(state="disabled")
            self.entry_lucro_por_unidade.delete(0, tk.END)  # Clear value when disabling


if __name__ == "__main__":
    root = tk.Tk()
    app = InterfaceProdutos(root)
    root.mainloop()

